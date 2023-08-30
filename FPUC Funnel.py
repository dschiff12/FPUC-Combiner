import tkinter as tk
from tkinter import filedialog
import os
import pandas as pd
from openpyxl import load_workbook


def process_files():
    global state, combined_cu_path, material_sheet_path

    if state == 0:
        combined_cu_path = filedialog.askopenfilename(title="Upload Combined Materials Here", 
                                                      filetypes=[("Excel files", "*.xlsx *.xls *.xlsm")])
        if not combined_cu_path:
            print("Combined Materials file not selected. Exiting.")
            status_label.config(text="Combined Materials file not selected.")
            return
        else:
            button.config(text="Upload File to Funnel Materials Into")
            state = 1
            return

    if state == 1:
        material_sheet_path = filedialog.askopenfilename(title="Upload File to Funnel Materials Into", 
                                                         filetypes=[("Excel files", "*.xlsx *.xls *.xlsm")])
        if not material_sheet_path:
            print("Material Sheet file not selected. Exiting.")
            status_label.config(text="Material Sheet file not selected.")
            return
        else:
            status_label.config(text="Processing Files...")
            root.update_idletasks()
    
    if not combined_cu_path or not material_sheet_path:
        print("Files not selected. Exiting.")
        status_label.config(text="Files not selected.")
        return
    
    status_label.config(text="Processing Files...")
    root.update_idletasks()

    # Load the Excel sheets
    combined_cu_data = pd.read_excel(combined_cu_path)
    material_sheet = load_workbook(material_sheet_path)

    # Define column offsets (0-based)
    part_no_column = 5  # Corresponds to column F
    quantity_column = 0 # Corresponds to column A

    # Your existing logic for iterating through the rows and sheets
    for index, row in combined_cu_data.iterrows():
        work_function = row['Work Function']
        part_no = row['PART NO.']
        quantity = row['Quantity']

        if work_function == 'I':
            for sheet_name in material_sheet.sheetnames:
                ws = material_sheet[sheet_name]
                for r_idx in range(1, 201):
                    if ws.cell(row=r_idx, column=part_no_column + 1).value == part_no:
                        current_quantity = ws.cell(row=r_idx, column=quantity_column + 1).value or 0
                        ws.cell(row=r_idx, column=quantity_column + 1, value=current_quantity + quantity)

    # Ask the user for the filename to save the workbook
    save_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                             filetypes=[("Excel files", "*.xlsx")],
                                             initialdir=os.path.dirname(material_sheet_path),
                                             title="Save the modified material sheet as")
    if not save_path:
        print("No save file selected. Exiting.")
        status_label.config(text="No save file selected.")
    else:
        material_sheet.save(save_path)
        print(f"Updated quantities successfully! Saved as {save_path}")
        status_label.config(text="Updated quantities successfully!")
        root.destroy()


state = 0
# Initialize Tkinter root window
root = tk.Tk()
root.title("Excel File Processor")
root.geometry("400x200")

# Add a button to trigger file selection and processing
button = tk.Button(root, text="Upload Combined Materials Here", command=process_files, height=3, width=50)
button.pack(pady=20)

# Add a label to display the status
status_label = tk.Label(root, text="", width=50)
status_label.pack(pady=10)

root.mainloop()
